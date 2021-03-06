#Pipeline Analysis Workflow
#Code generated by Loren Suite
#06/01/2021

#The Pipeline Analysis Workflow EmpowerIT file must be run first to extract the most updated tsp5_demand file.

#Import applicable libraries

#from win32com.constants import xlFillDefault
import win32com.client as win32
from win32com.client import DispatchEx
from win32com.client import DispatchEx
import openpyxl
import pandas as pd
import openpyxl
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
import random
from copy import copy
import time

from openpyxl.formula.translate import Translator

from win32com.client import constants

import struct
print(struct.calcsize("P")*8)

#Import library, record, and print the date of the workflow execution.
from datetime import date
today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)

#Creates variable for the necessary file destinations.
tsp5_file = 'C:\\Users\\Administrator\\Downloads\\tsp5_demand.xlsx'
PL_file = 'C:\\Users\\Administrator\\Downloads\\PL_20210521_Workflow.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
pl_template = 'C:\\Users\\Administrator\\Downloads\\template_file.xlsx'
pl_merged = 'C:\\Users\\Administrator\\Downloads\\PL_Merged.xlsx'
pl_template_rows = 'C:\\Users\\Administrator\\Downloads\\PL_rows.xlsx'
#Reads in first tab of tsp5_demand file as a DataFrame in pandas. Prints first 5 rows to verify.
df = pd.read_excel(tsp5_file)
print(df.head())

#op_nums = []
#for x in df['Opportunity Number']:
#    op_nums.append(int(x))

#Rename columns according to formula requirements.
df = df.rename(columns = {'Commit Level' : 'C/L BD', 'Contract Profit Percentage' : 'Fee %','Period of Performance(In Months)' : 'PoP'})

writer = pd.ExcelWriter(tsp5_updated)
df.to_excel(writer)
writer.save()

#Loads tsp5_updated into a workbook using Openpyxl.
ss = openpyxl.load_workbook(tsp5_updated)

#Opens the tsp5_demand updated file and changes the tab name to reflect the necessary formatting for the main file. Saves.
ss_sheet = ss['Sheet1']
ss_sheet.title = 'Data '+ newDate

#Deletes extra row tracking columns created in the updated DataFrame.
ss_sheet.delete_cols(1)
rowsNeeded = str(ss_sheet.max_row)
print(rowsNeeded)

#Defines the header name.
sheetid = ss.sheetnames.index('Data ' + newDate)
print(sheetid)
hdr_range = openpyxl.workbook.defined_name.DefinedName('PLS_Hdr_'+ newDate, attr_text = "'Data " + newDate + "'!$A$1:$AV$1")
ss.defined_names.append(hdr_range)

#Defines the data name.
data_range = openpyxl.workbook.defined_name.DefinedName('PLS_Data_'+ newDate, attr_text = "'Data " + newDate + "'!$A$1:$AV$" + rowsNeeded)
ss.defined_names.append(data_range)

#Manually set formatted column widths to match previous data files.
ss_sheet.column_dimensions['A'].width = 17.3
ss_sheet.column_dimensions['B'].width = 6
ss_sheet.column_dimensions['C'].width = 15.3
ss_sheet.column_dimensions['D'].width = 15.3
ss_sheet.column_dimensions['E'].width = 14.3
ss_sheet.column_dimensions['F'].width = 15.3
ss_sheet.column_dimensions['G'].width = 20.3
ss_sheet.column_dimensions['H'].width = 9.40
ss_sheet.column_dimensions['I'].width = 13.0
ss_sheet.column_dimensions['J'].width = 9.4
ss_sheet.column_dimensions['K'].width = 11.3
ss_sheet.column_dimensions['L'].width = 9.4
ss_sheet.column_dimensions['M'].width = 10.3
ss_sheet.column_dimensions['N'].width = 8.3
ss_sheet.column_dimensions['O'].width = 15.3
ss_sheet.column_dimensions['P'].width = 6.5
ss_sheet.column_dimensions['Q'].width = 17.3
ss_sheet.column_dimensions['R'].width = 9.3
ss_sheet.column_dimensions['S'].width = 14.2
ss_sheet.column_dimensions['T'].width = 17.3
ss_sheet.column_dimensions['U'].width = 11.3
ss_sheet.column_dimensions['V'].width = 17.3
ss_sheet.column_dimensions['W'].width = 19.3
ss_sheet.column_dimensions['X'].width = 16.30
ss_sheet.column_dimensions['Y'].width = 17.3
ss_sheet.column_dimensions['Z'].width = 15.3
ss_sheet.column_dimensions['AA'].width = 4
ss_sheet.column_dimensions['AB'].width = 10.3
ss_sheet.column_dimensions['AC'].width = 15.3
ss_sheet.column_dimensions['AD'].width = 19.3
ss_sheet.column_dimensions['AE'].width = 19.3
ss_sheet.column_dimensions['AF'].width = 19.3
ss_sheet.column_dimensions['AG'].width = 19.3
ss_sheet.column_dimensions['AH'].width = 19.3
ss_sheet.column_dimensions['AI'].width = 16.5
ss_sheet.column_dimensions['AJ'].width = 15.3
ss_sheet.column_dimensions['AK'].width = 12.3
ss_sheet.column_dimensions['AL'].width = 9.40
ss_sheet.column_dimensions['AM'].width = 15.3
ss_sheet.column_dimensions['AN'].width = 17.3
ss_sheet.column_dimensions['AO'].width = 17.3
ss_sheet.column_dimensions['AP'].width = 9.3
ss_sheet.column_dimensions['AQ'].width = 16
ss_sheet.column_dimensions['AR'].width = 16.3
ss_sheet.column_dimensions['AS'].width = 15.3
ss_sheet.column_dimensions['AT'].width = 24.3
ss_sheet.column_dimensions['AU'].width = 14.3
ss_sheet.column_dimensions['AV'].width = 14.3

#Format all cells in tsp5_demand to be text wrapped and with a top vertical adjustment.
for row in ss_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text = True,vertical = 'top')

#Saves the updated tsp5 file before merging with the PL template excel file.
ss.save(tsp5_updated)

#Opens the PL_merged file using Openpyxl as a workbook and creates a worksheet for each of the tabs.
pl_custom = openpyxl.load_workbook(pl_template)
#Creates a worksheet for the first tab containing the tsp5_demand data.
pl_temp = pl_custom.worksheets[0]
pl_temp.insert_cols(1)

#Creates arrays for headers and columns. Updates the PLS_Data and PLS_Hdr name ranges to reflect the new date of data
#import.
header_rows = []
column_formulas = []
for row_cells in pl_temp.iter_rows(min_row=16, max_row=16):
    for cell in row_cells:
        header = {}
        header['value'] = cell.value
        header['fill'] = cell.fill
        print(header['fill'])
        header_rows.append(cell)

for row_cells in pl_temp.iter_rows(min_row=17, max_row=17):
    for cell in row_cells:
        cell_dict = {}
        formula = cell.value

        if formula and isinstance(formula, str):
            formula = formula.replace('PLS_Data_Latest', 'PLS_Data_' + newDate)
            formula = formula.replace('PLS_Hdr_Latest', 'PLS_Hdr_' + newDate)
        cell_dict['formula'] = formula
        cell_dict['cell'] = cell
        column_formulas.append(cell_dict)

row = 16
for c in range(1, len(header_rows)):
    pl_temp.cell(row=row, column=c).value = copy(header_rows[c].value)
    pl_temp.cell(row=row, column=c).fill = copy(header_rows[c].fill)
    pl_temp.cell(row=row, column=c).number_format = copy(header_rows[c].number_format)

row = 17
for c in range(1, len(header_rows)):
    pl_temp.cell(row=row, column=c).value = copy(column_formulas[c]['formula'])
    pl_temp.cell(row=row, column=c).fill = copy(column_formulas[c]['cell'].fill)
    pl_temp.cell(row=row, column=c).number_format = copy(column_formulas[c]['cell'].number_format)

pl_custom.save(pl_template_rows)

import win32com.client as win32
excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = True
wt = excel.Workbooks.Open(pl_template_rows)
ws = wt.Worksheets('PL Template')
excel.AskToUpdateLinks = False
opp_row_count = 17

excel.CutCopyMode = False

#while opp_row_count < 25:
#    print(opp_row_count)
#    ws.Range('B'+ str(opp_row_count) +':J'+ str(opp_row_count)).Select()
#    excel.Selection.AutoFill(ws.Range('B'+str(opp_row_count)+':J'+str(opp_row_count+1),win32.constants.xlFillDefault))
#    wt.SaveAs(pl_template_rows)
#    time.sleep(2)
#    opp_row_count = opp_row_count + 1

ws.Range('B17:EJ17').Select()
excel.Selection.AutoFill(ws.Range('B17:EJ25'),win32.constants.xlFillDefault)
#excel.CutCopyMode = False

#time.sleep(2)
#ws.Range('F17').Select()
#excel.Selection.AutoFill(ws.Range('F17:F'+str(int(rowsNeeded))),win32.constants.xlFillDefault)
#excel.CutCopyMode = False

#ws.Range('G17').Select()
#excel.Selection.AutoFill(ws.Range('G17:G'+str(int(rowsNeeded))),win32.constants.xlFillDefault)
#excel.CutCopyMode = False

#ws.Range('B19:EK19').Select()
#excel.Selection.AutoFill(ws.Range('B19:EK20'),win32.constants.xlFillDefault)

#excel.DisplayAlerts = False
wt.SaveAs(pl_template_rows)
excel.Application.Quit()

#Opens the two workbooks needed and merges them by placing the single tab from tsp5_demand before the first tab of the
#main file. Saves the merged workbook as PL_Merged and opens the Excel file for viewing. Quits excel to prepare for next
#actions.
excel = DispatchEx('Excel.Application')
excel.Visible = False
PL_workbook = excel.Workbooks.Open(pl_template_rows)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
tsp5_workbook.Worksheets(ss_sheet.title).Move(Before = PL_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_workbook.SaveAs(pl_merged)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')

#Opens the PL_merged file using Openpyxl as a workbook and creates a worksheet for each of the tabs.
data_file = openpyxl.load_workbook(pl_merged)
#Creates a worksheet for the first tab containing the tsp5_demand data.
data_sheet = data_file.worksheets[0]
#Creates a worksheet for the second tab containing the PL_template sheet.
date_sheet = data_file.worksheets[1]
date_sheet.insert_cols(1)

#Creates a string holding the number of opportunity numbers found in the data sheet.
baseRows = data_sheet.max_row
strRows = str(baseRows)

maxCol = date_sheet.max_column
strCol = str(maxCol)
print(strCol + 'max cols')

#Renames the PL Template tab to reflect the date of data import
date_sheet.title = 'PL '+ newDate

#Defines copyRange to bring in a specified range of cells to be copied over for transfer.
def copyRange(startCol, startRow, endCol, endRow ,sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(data_sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
    return rangeSelected

#Define pasteRange to dictate where the cells should be pasted in the recieving document.
def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 1
    for i in range(0, len(copiedData), 1):
        countCol = 1
        for j in range(startCol, endCol+1, 1):
            sheetRecieving.cell(row = startRow + i, column = startCol,value = copiedData[i][0])
            countCol += 1
        countRow += 1

#Defines createData to copy and paste the data from one sheet to another. Copies the opportunity column from the tsp5
#data sheet and pastes it into the pre-formatted PL_template opportunity number column starting in column B and on row
#17. Prints at the beginning and end of the process and saves the file. New functions are created for each column that
#needs to be copied over.
def oppNumCopy():
    selectedRange = copyRange(1, 2, 1, baseRows, data_sheet)
    pastingRange = pasteRange(2, 17, 2, baseRows + 18,date_sheet, selectedRange)
    print('Opportunity Number range copied and pasted.')

oppNumCopy()

def rowCopy():
    selectedRange = copyRange(3, 17, maxCol - 2, 17, date_sheet)
    pastingRange = pasteRange(3, 18, maxCol - 2, 18, date_sheet, selectedRange)
    print('Opportunity Number range copied and pasted.')

rowCopy()

data_file.save(pl_merged)
print('All data has been copied to the PL tab successfully.')
excel.Application.Quit()

#excel = DispatchEx('Excel.Application')
#excel.Visible = False
#Merged_workbook = excel.Workbooks.Open(pl_merged)
#Original_workbook = excel.Workbooks.Open(PL_file, read_Only=True)
#Merged_workbook.Worksheets(Merged_workbook.Worksheets(2).title).Move(Before = Original_workbook.Worksheets(1))
#Merged_workbook.Worksheets(Merged_workbook.Worksheets(1).title).Move(Before = Original_workbook.Worksheets(1))
#excel.DisplayAlerts = False
#final_workbook.SaveAs(updated_file)
#excel.DisplayAlerts = True
#excel.Visible = True
#excel.Application.Quit()
#print('Files have been successfully merged.')

#date_sheet['C18'].formula = column_formulas[3]
#print(date_sheet['C18'].formula)

#rowCount = str(18)
#count = 0
#currentCell = 'B'+rowCount
#colNum = str(3)
#print(date_sheet['B'+ int(rowCount)].value)


#while int(colNum) <= maxCol:
#    while int(rowCount) <= int(baseRows + 15):
        #date_sheet.cell(row = int(rowCount), column = int(colNum)).value = copy(column_formulas[int(colNum)]['formula'])
#        date_sheet.cell(row=int(rowCount), column=6).value = copy(column_formulas[6]['formula'])
        #date_sheet.cell(row = int(rowCount), column=6).fill = copy(column_formulas[6]['cell'].fill)
        #date_sheet.cell(row=int(rowCount), column=6).number_format = copy(column_formulas[6]['cell'].number_format)
#        date_sheet.cell(row=int(rowCount), column=7).value = copy(column_formulas[7]['formula'])
#        date_sheet.cell(row=int(rowCount), column=7).fill = copy(column_formulas[7]['cell'].fill)
#        date_sheet.cell(row=int(rowCount), column=7).number_format = copy(column_formulas[7]['cell'].number_format)
#        dateUpdate = 'F'+str(rowCount)
#        date_sheet[dateUpdate].value = date_sheet[dateUpdate].value.replace('$B17', '$B'+ str(rowCount))
#        dateUpdate = 'G' + str(rowCount)
#        date_sheet[dateUpdate].value = date_sheet[dateUpdate].value.replace('$B17', '$B' + str(rowCount))
#        rowCount = int(rowCount) + 1
#        print(rowCount)
#        currentCell = 'B'+ str(rowCount)
#        colNum = int(colNum) + 1

#if date_sheet['B'+ str(rowCount)].value >> 1:
    #print(date_sheet[currentCell].value)

#while int(colNum) < maxCol:
    #rowCount = str(18)
    #while int(rowCount) < int(baseRows + 15):
        #date_sheet.cell(row=int(rowCount), column=(int(colNum))).value = copy('=LET(PLSdata,VLOOKUP($B17,PLS_Data_20210618,MATCH(F$16,PLS_Hdr_20210618,0),FALSE),IF(PLSdata=0,"",PLSdata))')
        #date_sheet.cell(row=int(rowCount), column=(int(colNum))).value = copy(column_formulas[int(colNum)]['formula'])
           # if formula and isinstance(formula, str):
              #  formula = formula.replace('$B17', '$B' + str(rowCount))
              #  cell_dict['formula'] = formula
              #  cell_dict['cell'] = cell
              #  column_formulas.append(cell_dict)

        #date_sheet.cell(row = int(rowCount), column=6).fill = copy(column_formulas[6]['cell'].fill)
        #date_sheet.cell(row=int(rowCount), column=6).number_format = copy(column_formulas[6]['cell'].number_format)
        #dateUpdate = 'AB' + str(rowCount)
        #date_sheet[dateUpdate].value = date_sheet[dateUpdate].value.replace('17', str(rowCount))

        #rowCount = int(rowCount) + 1
        #print(rowCount)
        #currentCell = 'B'+ str(rowCount)
    #colNum = int(colNum) + 1

#Opens the two workbooks needed and merges them by placing the single tab from tsp5_demand before the first tab of the
#main file. Saves the merged workbook as PL_Merged and opens the Excel file for viewing. Quits excel to prepare for next
#actions.

#opp_num_list = []
#for x in pl_merged['Opportunity Number']:
#    opp_num_list.append(int(x))

#Final file save. Close out of excel.
#excel.DisplayAlerts = False
#PL_workbook.SaveAs(updated_file)
#excel.DisplayAlerts = True


#excel = DispatchEx('Excel.Application')
#excel.Visible = False
#PL_workbook = excel.Workbooks.Open(pl_template)
#tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
#xlSheet_to_final.Range("B18:EJ18").AutoFill(xlSheet_to_final.Range('B19:EJ19"), xlFillDefault)

#ws.Range("C3:K3").Select()
#excel.Selection.AutoFill(ws.Range("C3:K11"),win32.constants.xlFillDefault)
#
# Autofill cell contents
#
#import win32com.client as win32
#excel = win32.gencache.EnsureDispatch('Excel.Application')
#wb = excel.Workbooks.Add()
#ws = wb.Worksheets("Sheet1")
#ws.Range("C17").Value = column_formulas[3]
#ws.Range("D17").Value = column_formulas[4]
#ws.Range("C17:D17").AutoFill(ws.Range("C17:D20"),win32.constants.xlFillDefault)
#excel.Application.Quit()
#from win32com.client.gencache import EnsureDispatch
#from win32com.client import constants

#def Macro1():
#    xl = EnsureDispatch('Excel.Application')
#    xl.Range('B11:K11').Select()
#    xl.Selection.AutoFill(Destination=xl.Range('B11:K16'), Type=constants.xlFillDefault)
#    xl.Columns('B:K').Select()
#    xl.Selection.ColumnWidth = 4

#Macro1()