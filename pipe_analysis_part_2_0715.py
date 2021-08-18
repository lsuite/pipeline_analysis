import openpyxl
import pathlib
import win32com.client as win32
import os

import logging
logging.basicConfig(filename='pipelineanalysis.log', level=logging.INFO)

#Creates variable for the necessary file destinations.
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
pl_template_rows = 'C:\\Users\\Administrator\\Downloads\\PL_rows.xlsx'
test_copy_merge = "C:\\Users\\Administrator\\Downloads\\copy_merge.xlsx"
PL_file = 'C:\\Users\\Administrator\\Downloads\\PL_20210521_Workflow.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
test_save = 'C:\\Users\\Administrator\\Downloads\\test_save.xlsx'
excel = win32.gencache.EnsureDispatch('Excel.Application')

row_file = openpyxl.load_workbook(tsp5_updated)
row_sheet = row_file.worksheets[0]
row_file.close()

maxRow = row_sheet.max_row + 15
strRow = str(maxRow)

#Merges the updated tsp5_demand file and the autofilled template file into the master Excel file as tabs.
excel.Visible = False
PL_template_workbook = excel.Workbooks.Open(pl_template_rows)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
tsp5_workbook.Worksheets(1).Move(Before = PL_template_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_template_workbook.SaveAs(test_copy_merge)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

# Import library, record, and print the date of the workflow execution.
from datetime import date

today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)


def get_cells_as_values(path_to_xlsx, sheetname, valuerange):
    """
    Open a workbook, select a sheet(sheetname) and return a
    dictionary of Cells where the key is the cell location (A10, for example)
    """

    # Open the workbook and get the appropriate worksheet
    workbook = openpyxl.load_workbook(path_to_xlsx, data_only=True)
    worksheet = workbook[sheetname]

    # Grabs the range of cells we're interested in:
    cells_subcolumn = worksheet[valuerange]
    return_dict = {}
    for cell_tuple in cells_subcolumn:
        for cell in cell_tuple: # Cell is initially a tuple; grab the first cell
            return_dict[cell.coordinate] = cell.value

    return return_dict


def main():
    # Generate a path to the xlsx file - relative to this file's location
    path = pathlib.Path(__file__).parent  # Grabs the file's folder
    xlsx_path = path / "C:\\Users\\Administrator\\Downloads\\copy_merge.xlsx"

    #Defines target cells ranges for copying and pasting as values and retrieves those cells as values.
    cell_range = "BB17:BK" + strRow
    cell_range2 = "F17:H" + strRow
    cell_range3 = "CU17:EA" + strRow
    sheet_name = "PL " + newDate
    cell_values = get_cells_as_values(xlsx_path, sheet_name, cell_range)
    cell_values2 = get_cells_as_values(xlsx_path, sheet_name, cell_range2)
    cell_values3 = get_cells_as_values(xlsx_path, sheet_name, cell_range3)

    # Reopen, but as normal:
    workbook = openpyxl.load_workbook(xlsx_path)
    worksheet = workbook[sheet_name]
    cells_with_formulae = worksheet[cell_range]
    cells_with_formulae2 = worksheet[cell_range2]
    cells_with_formulae3 = worksheet[cell_range3]

    # Iterate over cells; determine coordinate and grab value from cell_values dict;
    # Set cell to value for cell range 1. Prints confirmation statement.
    for cell_tuple in cells_with_formulae:
        for cell in cell_tuple:
            new_value = cell_values[cell.coordinate]

            cell.value = new_value
    print('Range ' + cell_range + ' pasted as values.')

    #Set cell to value for cell range 2. Prints confirmations statement.
    for cell_tuple in cells_with_formulae2:
        for cell in cell_tuple:
            new_value2 = cell_values2[cell.coordinate]

            cell.value = new_value2
    print('Range ' + cell_range2 + ' pasted as values.')

    # Set cell to value for cell range 3. Prints confirmations statement.
    for cell_tuple in cells_with_formulae3:
        for cell in cell_tuple:
            new_value3 = cell_values3[cell.coordinate]

            # print old and new value:
            cell.value = new_value3
    print('Range ' + cell_range3 + ' pasted as values.')

    # Save the workbook to the current file location.
    workbook.save(path / xlsx_path)
    print('Copy ranges and paste as values complete.')
    workbook.close()

if __name__ == "__main__": main()

excel.Application.Quit()

print("Test1")
copy_workbook = openpyxl.load_workbook(test_copy_merge)
del copy_workbook.defined_names['PLS_Hdr_' + newDate]
del copy_workbook.defined_names['PLS_Data_' + newDate]
print(copy_workbook.get_named_ranges())
del copy_workbook['Data '+newDate]
copy_workbook.save(test_copy_merge)
copy_workbook.close()
print("Test2")
copy_workbook = excel.Workbooks.Open(test_copy_merge)
print('Test')

# Merges the updated tsp5_demand file and the autofilled template file into the master Excel file as tabs.
excel.Visible = False
PL_workbook = excel.Workbooks.Open(PL_file)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
copy_workbook.Worksheets(1).Move(Before=PL_workbook.Worksheets(1))
tsp5_workbook.Worksheets(1).Move(Before=PL_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_workbook.SaveAs(updated_file)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

#Removes copies of intermediate files completed during the workflow.
#os.remove(tsp5_updated)
#os.remove(pl_template_rows)
#os.remove(test_copy_merge)

#Removed code, saved in case it is needed later
    # Iterate again, printing new values
    #for cell_tuple in cells_with_formulae:
    #    for cell in cell_tuple:
    #        print(cell.value)

    # Iterate again, printing new values
    #for cell_tuple in cells_with_formulae2:
    #    for cell in cell_tuple:
    #        print(cell.value)

    # Iterate again, printing new values
    #for cell_tuple in cells_with_formulae3:
    #    for cell in cell_tuple:
    #        print(cell.value)