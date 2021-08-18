import win32com.client as win32
from win32com.client import DispatchEx
from win32com.client import DispatchEx
import openpyxl
import pandas as pd
import openpyxl
import os
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
import random
from copy import copy
import time

from openpyxl.formula.translate import Translator

import logging
logging.basicConfig(filename='pipelineanalysis.log', level=logging.INFO)

from win32com.client import constants

logging.info('Test 2 File Started.')
#Import library, record, and print the date of the workflow execution.
from datetime import date
today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)

#Creates variable for the necessary file destinations.
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
pl_template_rows = 'C:\\Users\\Administrator\\Downloads\\PL_rows.xlsx'
test_copy_merge = 'C:\\Users\\Administrator\\Downloads\\copy_merge.xlsx'
data_only = 'C:\\Users\\Administrator\\Downloads\\data_only.xlsx'
PL_file = 'C:\\Users\\Administrator\\Downloads\\PL_20210521_Workflow.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'

excel = win32.gencache.EnsureDispatch('Excel.Application')

row_file = openpyxl.load_workbook(tsp5_updated)
row_sheet = row_file.worksheets[0]

hdrs = []
for cell in row_sheet[1]:
    hdrs.append(cell.value)
print(hdrs)

maxRow = row_sheet.max_row - 1
strRow = str(maxRow)

#Merges the updated tsp5_demand file and the autofilled template file into the master Excel file as tabs.
excel.Visible = False
PL_template_workbook = excel.Workbooks.Open(pl_template_rows)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
tsp5_workbook.Worksheets(1).Move(Before = PL_template_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_template_workbook.SaveAs(test_copy_merge)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

header_merge_file = openpyxl.load_workbook(test_copy_merge)
hdr_search = header_merge_file.worksheets[1]
maxCol = hdr_search.max_column
hdrs2 = []
indxs = []
matched_hdrs = []
j = 0
i = 0
for cell in hdr_search[16]:
    hdrs2.append(cell.value)
print(hdrs2)

while j < len(hdrs):
    while i < len(hdrs2):
        if hdrs[j] == hdrs2[i]:
            print(hdrs[j] + ' and ' + hdrs2[i] + ' are a match at column ' + str(i) + '.')
            indxs.append(i)
            matched_hdrs.append(hdrs[j])
        i = i + 1
    j = j + 1
    i = 0
print(indxs)
print(matched_hdrs)

header_merge_file.close()

data_only_file = openpyxl.load_workbook(test_copy_merge, data_only= True)
data_only_file.save(data_only)
data_only_sheet = data_only_file.worksheets[1]
data_only_file.close()

#Opens the updated tsp5  file using Openpyxl as a workbook and creates a worksheet for the tab.
main_PL_file = openpyxl.load_workbook(test_copy_merge)
data_sheet = main_PL_file.worksheets[0]
pl_sheet = main_PL_file.worksheets[1]

#Creates a string holding the number of opportunity numbers found in the data sheet.
baseRows = data_sheet.max_row
strRows = str(baseRows)

#Defines copyRange to bring in a specified range of cells to be copied over for transfer.
def copyRange(startCol, startRow, endCol, endRow ,sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(data_only_sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
    return rangeSelected

#Define pasteRange to dictate where the cells should be pasted in the recieving document.
def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 1
    for i in range(0, len(copiedData), 1):
        countCol = 1
        for j in range(startCol, endCol+1, 1):
            sheetRecieving.cell(row = startRow + i, column = startCol,value = copiedData[i][0])
            countCol += 1
        countRow += 1

#Defines createData to copy and paste the data from one sheet to another. Copies the opportunity column from the tsp5
#data sheet and pastes it into the pre-formatted PL_template opportunity number column starting in column B and on row
#17. Prints at the end of the process to confirm.
def oppNumCopy():
    i = 1
    while i < len(indxs):
        selectedRange = copyRange(indxs[i], 17, indxs[i], baseRows + 15, data_only_sheet)
        #print(selectedRange)
        pastingRange = pasteRange(indxs[i], 17, indxs[i], baseRows + 15, pl_sheet, selectedRange)
        #print(pastingRange)
        print('Opportunity Number range copied and pasted.')
        print(indxs[i])
        i = i + 1

selectedRange = copyRange(105, 17, 105, baseRows + 15, data_only_sheet)
pastingRange = pasteRange(105, 17, 105, baseRows + 15, pl_sheet, selectedRange)
print('Opportunity Number range copied and pasted.')
oppNumCopy()

main_PL_file.save(test_copy_merge)
main_PL_file.close()

excel.Visible = False
PL_workbook = excel.Workbooks.Open(PL_file)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
#copy_workbook = excel.Workbooks.Open(test_copy_merge)
#copy_workbook.Worksheets(2).Move(Before=PL_workbook.Worksheets(1))
#copy_workbook.Worksheets(1).Move(Before=PL_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_workbook.SaveAs(updated_file)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

#os.remove(data_only)