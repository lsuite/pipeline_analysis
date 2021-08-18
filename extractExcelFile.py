import openpyxl
from openpyxl import Workbook
from datetime import date

#Record date of new data export from ServiceNow
today = date.today()
newDate = today.strftime("%Y%d%m")
print("Today's date:", newDate)

#Access the newly downloaded Excel file from ServiceNow after running the EmpowerIT bot file
pathServiceNow = 'C:\\Users\\Administrator\\Downloads\\tsp5_demand.xlsx'

wb_obj_ServiceNow = openpyxl.load_workbook(pathServiceNow)
sheet_obj_ServiceNow = wb_obj_ServiceNow.active

cell_obj_ServiceNow = sheet_obj_ServiceNow.cell(row=3, column = 4)
serviceNowRows = sheet_obj_ServiceNow.max_row
serviceNowColumns = sheet_obj_ServiceNow.max_column

print('The most updated data collected on', newDate ,'from ServiceNow contains:', serviceNowRows ,'rows and', serviceNowColumns , 'columns.')

#Access the main Excel file for storing data
pathBase = 'C:\\Users\\Administrator\\Downloads\\PL_20210409_Workflow.xlsx'

#Opening the workbook as read only due to the size
wb_obj_Base = openpyxl.load_workbook(pathBase, read_only = True)
sheet_obj_Base = wb_obj_Base.active

#Creating a write-only copy of the main Excel file for appending updates
wb_writeOnly = Workbook(write_only = True)
ws = wb_writeOnly.create_sheet('Update')

cell_obj_Base = sheet_obj_Base.cell(row = 3, column = 4)
baseRows = sheet_obj_Base.max_row

print(baseRows)

#Rename the new PL and data tabs with the updated download date
baseSheetList = wb_obj_Base.sheetnames
baseSheetList[2] = 'PL '+ newDate
baseSheetList[3] = 'PL '+ newDate
print(baseSheetList)

for i in range(0,len(wb_obj_Base.sheetnames)):
    wb_obj_Base.sheetnames[i]="new name"
#wb_obj_Base.save('PL 20210409 Workflow.xlsx')

#Determine if new rows were added in the recent download
#testRow = baseSheetList[2]
#baseRowMax = wb_obj_Base[testRow].max_row
#print(baseRowMax)
#if (baseRowMax<baseRows):
#    print('New Data Entries Have Been Added.')
#else: print('Same Row Count')

#Manually closing the main Excel file, required when using the read only command
wb_obj_Base.close