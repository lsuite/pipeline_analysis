import os
import pandas as pd

import openpyxl
from openpyxl import workbook

from datetime import date

#Record date of new data export from ServiceNow
today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)

writer = pd.ExcelWriter('C:\\Users\\Administrator\\pipelineAnalysis\\PL_20210409_Workflow.xlsx', engine = 'xlsxwriter')
writer.save()
empowerit_data = 'C:\\Users\\Administrator\\Downloads\\tsp5_demand.xlsx'

new_data = pd.read_excel(empowerit_data)

new_data.to_excel(writer, sheet_name = 'Data '+ newDate, index = False)
writer.save()

print('The most updated data collected on', newDate ,'from ServiceNow contains:', new_data.shape ,'rows and' , 'columns.')

print(new_data.head())

base_location = 'C:\\Users\\Administrator\\pipelineAnalysis\\PL_20210409_Workflow.xlsx'
base_file = pd.read_excel(base_location,sheet_name = 1,index_col = 3)

print(base_file.head())

#baseSheetList = base_file.sheetnames
#baseSheetList[2] = 'PL '+ newDate
#baseSheetList[3] = 'PL '+ newDate
#print(baseSheetList)

#for i in range(0,len(base_file.sheetnames)):
#    base_file.sheetnames[i]="new name"
#base_file.save('PL 20210409 Workflow.xlsx')