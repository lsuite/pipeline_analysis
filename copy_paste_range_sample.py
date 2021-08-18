import win32com.client as win32
from win32com.client import DispatchEx
from win32com.client import DispatchEx
import openpyxl
import pandas as pd
import openpyxl
import os
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
import random
from copy import copy
import time

from openpyxl.formula.translate import Translator

import logging
logging.basicConfig(filename='pipelineanalysis.log', level=logging.INFO)

from win32com.client import constants

logging.info('Test 2 File Started.')
#Import library, record, and print the date of the workflow execution.
from datetime import date
today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)

#Creates variable for the necessary file destinations.
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
pl_template_rows = 'C:\\Users\\Administrator\\Downloads\\PL_rows.xlsx'
test_copy_merge = 'C:\\Users\\Administrator\\Downloads\\copy_merge.xlsx'
data_only = 'C:\\Users\\Administrator\\Downloads\\data_only.xlsx'
PL_file = 'C:\\Users\\Administrator\\Downloads\\PL_20210521_Workflow.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'

#os.remove(data_only)
#os.remove(test_copy_merge)
#os.remove(updated_file)

excel = win32.gencache.EnsureDispatch('Excel.Application')

row_file = openpyxl.load_workbook(tsp5_updated)
row_sheet = row_file.worksheets[0]

hdrs = []
for cell in row_sheet[1]:
    hdrs.append(cell.value)
print(hdrs)

maxRow = row_sheet.max_row - 1
pl_max_row = maxRow + 15

strRow = str(maxRow)
str_pl_max = str(pl_max_row)

pl_rows_def_names = openpyxl.load_workbook(pl_template_rows)


pl_rows_def_names.save(pl_template_rows)
pl_rows_def_names.close()

#Merges the updated tsp5_demand file and the autofilled template file into the master Excel file as tabs.
excel.Visible = False
PL_template_workbook = excel.Workbooks.Open(pl_template_rows)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
tsp5_workbook.Worksheets(1).Move(Before = PL_template_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_template_workbook.SaveAs(test_copy_merge)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

header_merge_file = openpyxl.load_workbook(test_copy_merge)
hdr_search = header_merge_file.worksheets[1]
maxCol = hdr_search.max_column

hdrs2 = []
indxs = []
data_indxs = []
matched_hdrs = []
j = 0
i = 0
for cell in hdr_search[16]:
    hdrs2.append(cell.value)
print(hdrs2)

while j < len(hdrs):
    while i < len(hdrs2):
        if hdrs[j] == hdrs2[i]:
            print(hdrs[j] + ' and ' + hdrs2[i] + ' are a match at column ' + str(i) + '.')
            data_indxs.append(j)
            indxs.append(i)
            matched_hdrs.append(hdrs[j])
        i = i + 1
    j = j + 1
    i = 0
print(indxs)
print(matched_hdrs)

header_merge_file.close()

data_only_file = openpyxl.load_workbook(test_copy_merge, data_only = True)
data_only_file.save(data_only)
data_only_sheet = data_only_file.worksheets[1]
data_only_file.close()

#Opens the updated tsp5  file using Openpyxl as a workbook and creates a worksheet for the tab.
main_PL_file = openpyxl.load_workbook(test_copy_merge)
data_sheet = main_PL_file.worksheets[0]
pl_sheet = main_PL_file.worksheets[1]

#Creates a string holding the number of opportunity numbers found in the data sheet.
baseRows = data_sheet.max_row
strRows = str(baseRows)

#Defines copyRange to bring in a specified range of cells to be copied over for transfer.
def copyRange(startCol, startRow, endCol, endRow ,sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(data_only_sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
    return rangeSelected

#Define pasteRange to dictate where the cells should be pasted in the recieving document.
def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 1
    for i in range(0, len(copiedData), 1):
        countCol = 1
        for j in range(startCol, endCol+1, 1):
            sheetRecieving.cell(row = startRow + i, column = startCol,value = copiedData[i][0])
            countCol += 1
        countRow += 1

#Defines createData to copy and paste the data from one sheet to another. Copies the opportunity column from the tsp5
#data sheet and pastes it into the pre-formatted PL_template opportunity number column starting in column B and on row
#17. Prints at the end of the process to confirm.
def oppNumCopy():
    i = 0
    while i < len(indxs):
        #selectedRange = copyRange(indxs[i], 17, indxs[i], baseRows + 15, data_only)
        selectedRange = copyRange(indxs[i], 17, indxs[i], baseRows + 15, data_only_sheet)
        pastingRange = pasteRange(indxs[i], 17, indxs[i], baseRows + 15, pl_sheet, selectedRange)
        #print('Opportunity Number range copied and pasted.')
        print('Copying ' + hdrs[data_indxs[i]] + ' to ' + hdrs2[indxs[i]])
        #print(indxs[i])
        i = i + 1

oppNumCopy()

main_PL_file.save(test_copy_merge)

#Defines the header name.
pl_hdr_range = openpyxl.workbook.defined_name.DefinedName('PL_Hdr_Latest', attr_text = "'PL " + newDate + "'!$B$16:$EO$16")
#main_PL_file.defined_names.append(pl_hdr_range)

#Defines the data name.
pl_data_range = openpyxl.workbook.defined_name.DefinedName('PL_Data_Latest', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
#main_PL_file.defined_names.append(pl_data_range)

plz_inwork_group_range = openpyxl.workbook.defined_name.DefinedName('PLZ_InWork_Group', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
#main_PL_file.defined_names.append(plz_inwork_group_range)

plz_othergroup_a_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_A', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
#main_PL_file.defined_names.append(plz_othergroup_a_range)

plz_othergroup_ar_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_AR', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
#main_PL_file.defined_names.append(plz_othergroup_ar_range)

plz_othergroup_b_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_B', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
#main_PL_file.defined_names.append(plz_othergroup_b_range)

pl_oppno_latest_range = openpyxl.workbook.defined_name.DefinedName('PL_Oppno_Latest', attr_text = "'PL " + newDate + "'!$B$16:$B$" + str_pl_max)
#main_PL_file.defined_names.append(pl_oppno_latest_range)

pl_rowno_latest_range = openpyxl.workbook.defined_name.DefinedName('PL_RowNo_Latest', attr_text = "'PL " + newDate + "'!$B$17:$B$" + str_pl_max)
#main_PL_file.defined_names.append(pl_rowno_latest_range)

#final_workbook = openpyxl.load_workbook(updated_file, read_only = True)
#print('Done.')

#Open file using win32
#excel.Visible = True
#final_workbook = excel.Workbooks.Open(updated_file)

#Open file using pandas
#final_workbook = pd.read_excel(updated_file)

#Open file using openpyxl
#final_workbook = openpyxl.load_workbook(updated_file, read_only=True)
#final_workbook.remove('PL 20210430')
#final_workbook.save(updated_file)
#final_workbook.close()

#os.remove(data_only)
#os.remove(test_copy_merge)
