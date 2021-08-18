import openpyxl
from openpyxl import Workbook

# opening the source excel file
filename = 'C:\\Users\\Administrator\\Downloads\\tsp5_demand.xlsx'

wb1 = openpyxl.load_workbook(filename)
ws1 = wb1['Page 1']

# opening the destination excel file
filename1 = 'C:\\Users\\Administrator\\Downloads\\Clone.xlsx'
wb2 = openpyxl.load_workbook(filename1)
ws22 = wb2['Sheet1']
ws2 = wb2.active

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row=i, column=j)

        # writing the read value to destination excel file
        ws22.cell(row=i, column=j).value = c.value

for i in range(1, mr + 1):
    for j in range(1, mc + 1):
        # reading cell value from source excel file
        c = ws22.cell(row=i, column=j)
        print(c.value)

# saving the destination excel file
wb2.save(str(filename1))