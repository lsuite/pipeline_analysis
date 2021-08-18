#Pipeline Analysis Workflow
#Code generated by Loren Suite
#06/01/2021

#The Pipeline Analysis Workflow EmpowerIT file must be run first to extract the most updated tsp5_demand file.

#Import applicable libraries

import win32com.client as win32
from win32com.client import DispatchEx
import openpyxl
import pandas as pd
import openpyxl
from openpyxl import workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell import Cell
import random
from copy import copy
import time
import pynput
from pynput.keyboard import Key, Controller
keyboard = Controller()

import logging
logging.basicConfig(filename='pipelineanalysis.log', level=logging.INFO)

from openpyxl.formula.translate import Translator

from win32com.client import constants

logging.info('Script Initialized.')
#Import library, record, and print the date of the workflow execution.
from datetime import date
today = date.today()
newDate = today.strftime("%Y%m%d")
print("Today's date:", newDate)


#Creates variable for the necessary file destinations.
tsp5_file = 'C:\\Users\\Administrator\\Downloads\\tsp5_demand.xlsx'
PL_file = 'C:\\Users\\Administrator\\Downloads\\PL 20210716 Workflow.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'
tsp5_updated = 'C:\\Users\\Administrator\\Downloads\\tsp5_updated.xlsx'
pl_template = 'C:\\Users\\Administrator\\Downloads\\No_Named_Ranges_Template.xlsx'
pl_merged = 'C:\\Users\\Administrator\\Downloads\\PL_Merged.xlsx'
pl_template_rows = 'C:\\Users\\Administrator\\Downloads\\PL_rows.xlsx'
test_copy_merge = 'C:\\Users\\Administrator\\Downloads\\copy_merge.xlsx'
data_only = 'C:\\Users\\Administrator\\Downloads\\data_only.xlsx'
updated_file = 'C:\\Users\\Administrator\\Downloads\\PL_Updated_Workflow.xlsx'

#Reads in first tab of tsp5_demand file as a DataFrame in pandas. Prints first 5 rows to verify.
df = pd.read_excel(tsp5_file)
print(df.head())

#Rename columns according to formula requirements.
df = df.rename(columns = {'Commit Level' : 'C/L BD', 'Contract Profit Percentage' : 'Fee %','Period of Performance(In Months)' : 'PoP'})

#Writes the changes from the imported tsp5_demand file to a new file called updated and saves.
writer = pd.ExcelWriter(tsp5_updated)
df.to_excel(writer)
writer.save()

#Loads tsp5_updated into a workbook using Openpyxl.
ss = openpyxl.load_workbook(tsp5_updated)

#Opens the tsp5_demand updated file and changes the tab name to reflect the necessary formatting for the main file. Saves.
ss_sheet = ss['Sheet1']
ss_sheet.title = 'Data '+ newDate

#Deletes extra row tracking columns created in the updated DataFrame.
ss_sheet.delete_cols(1)

#Reads in the number of rows needed or the number of opportunity numbers as a string.
rowsNeeded = str(ss_sheet.max_row)
print(rowsNeeded)

#Defines the header name.
sheetid = ss.sheetnames.index('Data ' + newDate)
print(sheetid)
hdr_range = openpyxl.workbook.defined_name.DefinedName('PLS_Hdr_'+ newDate, attr_text = "'Data " + newDate + "'!$A$1:$AV$1")
ss.defined_names.append(hdr_range)

#Defines the data name.
data_range = openpyxl.workbook.defined_name.DefinedName('PLS_Data_'+ newDate, attr_text = "'Data " + newDate + "'!$A$1:$AV$" + rowsNeeded)
ss.defined_names.append(data_range)

#Manually set formatted column widths to match previous data files.
ss_sheet.column_dimensions['A'].width = 17.3
ss_sheet.column_dimensions['B'].width = 6
ss_sheet.column_dimensions['C'].width = 15.3
ss_sheet.column_dimensions['D'].width = 15.3
ss_sheet.column_dimensions['E'].width = 14.3
ss_sheet.column_dimensions['F'].width = 15.3
ss_sheet.column_dimensions['G'].width = 20.3
ss_sheet.column_dimensions['H'].width = 9.40
ss_sheet.column_dimensions['I'].width = 13.0
ss_sheet.column_dimensions['J'].width = 9.4
ss_sheet.column_dimensions['K'].width = 11.3
ss_sheet.column_dimensions['L'].width = 9.4
ss_sheet.column_dimensions['M'].width = 10.3
ss_sheet.column_dimensions['N'].width = 8.3
ss_sheet.column_dimensions['O'].width = 15.3
ss_sheet.column_dimensions['P'].width = 6.5
ss_sheet.column_dimensions['Q'].width = 17.3
ss_sheet.column_dimensions['R'].width = 9.3
ss_sheet.column_dimensions['S'].width = 14.2
ss_sheet.column_dimensions['T'].width = 17.3
ss_sheet.column_dimensions['U'].width = 11.3
ss_sheet.column_dimensions['V'].width = 17.3
ss_sheet.column_dimensions['W'].width = 19.3
ss_sheet.column_dimensions['X'].width = 16.30
ss_sheet.column_dimensions['Y'].width = 17.3
ss_sheet.column_dimensions['Z'].width = 15.3
ss_sheet.column_dimensions['AA'].width = 4
ss_sheet.column_dimensions['AB'].width = 10.3
ss_sheet.column_dimensions['AC'].width = 15.3
ss_sheet.column_dimensions['AD'].width = 19.3
ss_sheet.column_dimensions['AE'].width = 19.3
ss_sheet.column_dimensions['AF'].width = 19.3
ss_sheet.column_dimensions['AG'].width = 19.3
ss_sheet.column_dimensions['AH'].width = 19.3
ss_sheet.column_dimensions['AI'].width = 16.5
ss_sheet.column_dimensions['AJ'].width = 15.3
ss_sheet.column_dimensions['AK'].width = 12.3
ss_sheet.column_dimensions['AL'].width = 9.40
ss_sheet.column_dimensions['AM'].width = 15.3
ss_sheet.column_dimensions['AN'].width = 17.3
ss_sheet.column_dimensions['AO'].width = 17.3
ss_sheet.column_dimensions['AP'].width = 9.3
ss_sheet.column_dimensions['AQ'].width = 16
ss_sheet.column_dimensions['AR'].width = 16.3
ss_sheet.column_dimensions['AS'].width = 15.3
ss_sheet.column_dimensions['AT'].width = 24.3
ss_sheet.column_dimensions['AU'].width = 14.3
ss_sheet.column_dimensions['AV'].width = 14.3

#Format all cells in tsp5_demand to be text wrapped and with a top vertical adjustment.
for row in ss_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text = True,vertical = 'top')

#Saves the updated tsp5 file before merging with the PL template excel file.
ss.save(tsp5_updated)

#Prints confirmation that the data file changes and formatting are complete.
print('Data file is updated.')
logging.info('TSP5_Demand File Updated.')

#Dispatches the Excel application for future commands.
excel = win32.gencache.EnsureDispatch('Excel.Application')

#Opens the PL template file using Openpyxl as a workbook and creates a worksheet for the tab.
template_file = openpyxl.load_workbook(pl_template)
print('Test')
template_sheet = template_file.worksheets[0]

#Records the maximum number of columns in the PL Template file as a string.
maxCol = template_sheet.max_column
strCol = str(maxCol)
print(strCol + 'max cols')

#Renames the PL Template tab to reflect the date of data import.
template_sheet.title = 'PL '+ newDate

#Opens the updated tsp5  file using Openpyxl as a workbook and creates a worksheet for the tab.
data_file = openpyxl.load_workbook(tsp5_updated)
data_sheet = data_file.worksheets[0]

#Creates a string holding the number of opportunity numbers found in the data sheet.
baseRows = data_sheet.max_row
strRows = str(baseRows)

#Defines copyRange to bring in a specified range of cells to be copied over for transfer.
def copyRange(startCol, startRow, endCol, endRow ,sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(data_sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
    return rangeSelected

#Define pasteRange to dictate where the cells should be pasted in the recieving document.
def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 1
    for i in range(0, len(copiedData), 1):
        countCol = 1
        for j in range(startCol, endCol+1, 1):
            sheetRecieving.cell(row = startRow + i, column = startCol,value = copiedData[i][0])
            countCol += 1
        countRow += 1

#Defines createData to copy and paste the data from one sheet to another. Copies the opportunity column from the tsp5
#data sheet and pastes it into the pre-formatted PL_template opportunity number column starting in column B and on row
#17. Prints at the end of the process to confirm.
def oppNumCopy():
    selectedRange = copyRange(1, 2, 1, baseRows, data_sheet)
    pastingRange = pasteRange(2, 17, 2, baseRows + 18,template_sheet, selectedRange)
    print('Opportunity Number range copied and pasted.')

oppNumCopy()
logging.info('Opportunity Number Copied to PL Tab')
#Takes in the headers (column names and sample data row) and pastes them back into the corresponding cells. Updates the
#PLS_Data/Hdr_Latest components of the formulas in row 17 and updates them to reflect the Define Names created above as
#PLS_Data/Hdr_NewDate.
header_rows = []
column_formulas = []
for row_cells in template_sheet.iter_rows(min_row=16, max_row=16):
    for cell in row_cells:
        header = {}
        header['value'] = cell.value
        header['fill'] = cell.fill
        print(header['fill'])
        header_rows.append(cell)

for row_cells in template_sheet.iter_rows(min_row=17, max_row=17):
    for cell in row_cells:
        cell_dict = {}
        formula = cell.value

        if formula and isinstance(formula, str):
            formula = formula.replace('PLS_Data_Latest', 'PLS_Data_' + newDate)
            formula = formula.replace('PLS_Hdr_Latest', 'PLS_Hdr_' + newDate)
        cell_dict['formula'] = formula
        cell_dict['cell'] = cell
        column_formulas.append(cell_dict)

row = 16
for c in range(1, len(header_rows)):
    template_sheet.cell(row=row, column=c+1).value = copy(header_rows[c].value)
    template_sheet.cell(row=row, column=c+1).fill = copy(header_rows[c].fill)
    template_sheet.cell(row=row, column=c+1).number_format = copy(header_rows[c].number_format)

row = 17
for c in range(1, len(header_rows)):
    template_sheet.cell(row=row, column=c+1).value = copy(column_formulas[c]['formula'])
    template_sheet.cell(row=row, column=c+1).fill = copy(column_formulas[c]['cell'].fill)
    template_sheet.cell(row=row, column=c+1).number_format = copy(column_formulas[c]['cell'].number_format)

#Saves the updated template file as PL_Template Rows and quits the Excel application.
template_file.save(pl_template_rows)
excel.Application.Quit()

time.sleep(10)
logging.info('Authofill Initialized.')
#Autofills the formulas to cover the number of columns needed based on the quantity of opportunity numbers.
excel.Visible = True
wt = excel.Workbooks.Open(pl_template_rows)
ws = wt.Worksheets[1]
excel.AskToUpdateLinks = False

ws.Range('C17:EO17').Select()
excel.Selection.AutoFill(ws.Range('C17:EO'+ str(baseRows+15)),win32.constants.xlFillDefault)
wt.Save()
excel.Application.Quit()

excel.DisplayAlerts = False
excel.AskToUpdateLinks = False

#Open tsp5_demand updated file and appends all of the headers to an array called hdrs.
row_file = openpyxl.load_workbook(tsp5_updated)
row_sheet = row_file.worksheets[0]

hdrs = []
for cell in row_sheet[1]:
    hdrs.append(cell.value)
print(hdrs)

#Take in the max rows from the data tab and creates string values for the total # of Rows and the total number of rows
#needed for the PL tab that starts on row 16. ---> Step can be consolidated later.
maxRow = row_sheet.max_row - 1
pl_max_row = maxRow + 15

strRow = str(maxRow)
str_pl_max = str(pl_max_row)

#Merges the updated tsp5_demand file and autofilled template file into one workbook to generate a complete PL tab.
excel.Visible = False
PL_template_workbook = excel.Workbooks.Open(pl_template_rows)
tsp5_workbook = excel.Workbooks.Open(tsp5_updated)
tsp5_workbook.Worksheets(1).Move(Before = PL_template_workbook.Worksheets(1))
excel.DisplayAlerts = False
PL_template_workbook.SaveAs(test_copy_merge)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

#Opens the merged Data and PL tab file, records the maximum number of columns.
header_merge_file = openpyxl.load_workbook(test_copy_merge)
hdr_search = header_merge_file.worksheets[1]
maxCol = hdr_search.max_column

#Stores the headers from the PL tab into an array called hdrs2[] and indexes which column the headers from the Data
#and PL tab match. Indexes are stored in an array called indxs[].
hdrs2 = []
indxs = []
data_indxs = []
matched_hdrs = []
j = 0
i = 0
for cell in hdr_search[16]:
    hdrs2.append(cell.value)
print(hdrs2)

while j < len(hdrs):
    while i < len(hdrs2):
        if hdrs[j] == hdrs2[i]:
            print(hdrs[j] + ' and ' + hdrs2[i] + ' are a match at column ' + str(i) + '.')
            data_indxs.append(j)
            indxs.append(i)
            matched_hdrs.append(hdrs[j])
        i = i + 1
    j = j + 1
    i = 0
print(indxs)
print(matched_hdrs)

header_merge_file.close()

#Creates a data only version of the PL tab in the Data and PL tab merged file.
data_only_file = openpyxl.load_workbook(test_copy_merge, data_only = True)
data_only_file.save(data_only)
data_only_sheet = data_only_file.worksheets[1]
data_only_file.close()

#Opens the updated tsp5  file using Openpyxl as a workbook and creates a worksheet for the tab.
main_PL_file = openpyxl.load_workbook(test_copy_merge)
data_sheet = main_PL_file.worksheets[0]
pl_sheet = main_PL_file.worksheets[1]

#Defines copyRange to bring in a specified range of cells to be copied over for transfer.
def copyRange(startCol, startRow, endCol, endRow ,sheet):
    rangeSelected=[]
    for i in range(startRow, endRow + 1, 1):
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(data_only_sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
    return rangeSelected

#Define pasteRange to dictate where the cells should be pasted in the recieving document.
def pasteRange(startCol, startRow, endCol, endRow, sheetRecieving, copiedData):
    countRow = 1
    for i in range(0, len(copiedData), 1):
        countCol = 1
        for j in range(startCol, endCol+1, 1):
            sheetRecieving.cell(row = startRow + i, column = startCol,value = copiedData[i][0])
            countCol += 1
        countRow += 1

#Defines createData to copy and paste the data from one sheet to another. Copies the opportunity column from the tsp5
#data sheet and pastes it into the pre-formatted PL_template opportunity number column starting in column B and on row
#17. Prints at the end of the process to confirm.
def oppNumCopy():
    i = 0
    while i < len(indxs):
        selectedRange = copyRange(indxs[i], 17, indxs[i], pl_max_row, data_only_sheet)
        pastingRange = pasteRange(indxs[i], 17, indxs[i], pl_max_row, pl_sheet, selectedRange)
        print('Copying ' + hdrs[data_indxs[i]] + ' to ' + hdrs2[indxs[i]])
        i = i + 1

oppNumCopy()

#Defines the header name.
pl_hdr_range = openpyxl.workbook.defined_name.DefinedName('PL_Hdr_Latest', attr_text = "'PL " + newDate + "'!$B$16:$EO$16")
main_PL_file.defined_names.append(pl_hdr_range)

#Defines the data name.
pl_data_range = openpyxl.workbook.defined_name.DefinedName('PL_Data_Latest', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
main_PL_file.defined_names.append(pl_data_range)

plz_inwork_group_range = openpyxl.workbook.defined_name.DefinedName('PLZ_InWork_Group', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
main_PL_file.defined_names.append(plz_inwork_group_range)

plz_othergroup_a_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_A', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
main_PL_file.defined_names.append(plz_othergroup_a_range)

plz_othergroup_ar_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_AR', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
main_PL_file.defined_names.append(plz_othergroup_ar_range)

plz_othergroup_b_range = openpyxl.workbook.defined_name.DefinedName('PLZ_OtherGroup_B', attr_text = "'PL " + newDate + "'!$B$16:$EO$" + str_pl_max)
main_PL_file.defined_names.append(plz_othergroup_b_range)

pl_oppno_latest_range = openpyxl.workbook.defined_name.DefinedName('PL_Oppno_Latest', attr_text = "'PL " + newDate + "'!$B$16:$B$" + str_pl_max)
main_PL_file.defined_names.append(pl_oppno_latest_range)

pl_rowno_latest_range = openpyxl.workbook.defined_name.DefinedName('PL_RowNo_Latest', attr_text = "'PL " + newDate + "'!$B$17:$B$" + str_pl_max)
main_PL_file.defined_names.append(pl_rowno_latest_range)

busunit_range = openpyxl.workbook.defined_name.DefinedName('BusUnit', attr_text = "'Lookup" + newDate + "'!$N$54:$O$64")
#main_PL_file.defined_names.append(busunit_range)

number_rows = pl_sheet.max_row
number_columns = pl_sheet.max_column

replacement = {'#N/A':' '}

for i in range(number_columns):
    for k in range(number_rows):
        cell = str(pl_sheet[get_column_letter(i+1)+str(k+1)].value)
        for key in replacement.keys():
            if str(cell) == key:
                newCell = replacement.get(key)
                pl_sheet[get_column_letter(i+1)+str(k+1)] = str(newCell)

main_PL_file.save(test_copy_merge)
main_PL_file.close()

excel.Visible = True
PL_workbook = excel.Workbooks.Open(PL_file)
time.sleep(10)
copy_workbook = excel.Workbooks.Open(test_copy_merge)
tsp5_new = excel.Workbooks.Open(tsp5_updated)
print('Test')
time.sleep(10)
excel.DisplayAlerts = False
copy_workbook.Worksheets(2).Move(Before=PL_workbook.Worksheets(1))
time.sleep(10)
print('Test')
tsp5_new.Worksheets(1).Move(Before=PL_workbook.Worksheets(1))
time.sleep(10)
excel.DisplayAlerts = False
PL_workbook.SaveAs(updated_file)
excel.DisplayAlerts = True
excel.Visible = True
excel.Application.Quit()
print('Files have been successfully merged.')
logging.info('Workbooks Merged into Final Document')

#Autofills the formulas to cover the number of columns needed based on the quantity of opportunity numbers.
excel.Visible = True
wt = excel.Workbooks.Open(updated_file)
ws = wt.Worksheets['PL '+newDate]
excel.AskToUpdateLinks= False

ws.Range('C17').Select()
excel.Selection.AutoFill(ws.Range('C17:C'+ str(baseRows+15)),win32.constants.xlFillDefault)
wt.Save()
excel.Application.Quit()

excel.DisplayAlerts = False
excel.AskToUpdateLinks = False
#test = openpyxl.load_workbook(updated_file, read_only = True)

